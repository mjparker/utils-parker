import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
import os
import sys


def format_csv_to_excel(csv_file_path, xlsx_file_path=None):
    """
    Convert CSV to formatted XLSX with:
    - Bold first row
    - AutoFilter enabled
    - Top row frozen
    """

    # Generate output filename if not provided
    if xlsx_file_path is None:
        base_name = os.path.splitext(csv_file_path)[0]
        xlsx_file_path = f"{base_name}.xlsx"

    try:
        # Read CSV file
        print(f"Reading CSV file: {csv_file_path}")
        df = pd.read_csv(csv_file_path)

        # Save as Excel file
        print("Converting to Excel format...")
        df.to_excel(xlsx_file_path, index=False, sheet_name="Data")

        # Load the workbook to apply formatting
        print("Applying formatting...")
        wb = load_workbook(xlsx_file_path)
        ws = wb.active

        # Bold the first row (headers)
        bold_font = Font(bold=True)
        for cell in ws[1]:  # First row
            cell.font = bold_font

        # Enable AutoFilter
        ws.auto_filter.ref = ws.dimensions

        # Freeze the top row
        ws.freeze_panes = "A2"

        # Auto-adjust column widths (optional)
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (TypeError, AttributeError):
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save the formatted workbook
        wb.save(xlsx_file_path)
        print(f"Successfully created formatted Excel file: {xlsx_file_path}")

        return xlsx_file_path

    except Exception as e:
        print(f"Error processing file: {str(e)}")
        return None


def batch_convert_csvs(folder_path):
    """
    Convert all CSV files in a folder to formatted XLSX files
    """
    csv_files = [f for f in os.listdir(folder_path) if f.lower().endswith(".csv")]

    if not csv_files:
        print("No CSV files found in the specified folder.")
        return

    print(f"Found {len(csv_files)} CSV files to convert:")

    for csv_file in csv_files:
        csv_path = os.path.join(folder_path, csv_file)
        xlsx_path = os.path.join(folder_path, csv_file.replace(".csv", ".xlsx"))

        print(f"\nProcessing: {csv_file}")
        result = format_csv_to_excel(csv_path, xlsx_path)

        if result:
            print(f"✓ Converted: {csv_file} → {os.path.basename(result)}")
        else:
            print(f"✗ Failed to convert: {csv_file}")


if __name__ == "__main__":
    if len(sys.argv) == 2:
        input_path = sys.argv[1]

        if os.path.isfile(input_path) and input_path.lower().endswith(".csv"):
            # Single file conversion
            format_csv_to_excel(input_path)
        elif os.path.isdir(input_path):
            # Batch conversion of folder
            batch_convert_csvs(input_path)
        else:
            print("Please provide a valid CSV file or folder path.")
    else:
        print("Usage:")
        print("  Single file: python script.py path/to/file.csv")
        print("  Batch folder: python script.py path/to/folder/")
        print("\nOr use the functions directly:")
        print("  format_csv_to_excel('input.csv', 'output.xlsx')")
        print("  batch_convert_csvs('/path/to/csv/folder/')")
